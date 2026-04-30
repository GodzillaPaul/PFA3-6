"""Convert Excel print page to a faithful HTML rendition.

Reads `source.xlsx` (LibreOffice-converted .xls), extracts every cell of the
visible 列印頁 sheet with its merged-cell layout, fonts, alignment, fill,
borders and column widths, then emits a single static HTML page that mirrors
the spreadsheet pixel-for-pixel for online viewing.
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import html
import json
import re

SRC = 'source.xlsx'
OUT = 'index.html'
SHEET = '列印頁'

# Excel default column width unit -> approximate pixels (8.43 ~= 64px)
def col_px(width):
    if not width:
        return 56  # default ~ 8.43
    return max(1, int(round(width * 7 + 5)))

def row_px(height):
    if not height:
        return 18
    return max(1, int(round(height * 96 / 72)))  # pt to px

def color_to_hex(c):
    if c is None:
        return None
    try:
        if c.type == 'rgb' and c.rgb and isinstance(c.rgb, str):
            v = c.rgb
            if len(v) == 8:  # ARGB
                return '#' + v[2:]
            if len(v) == 6:
                return '#' + v
        # theme/indexed -> we ignore (default)
    except Exception:
        return None
    return None

def fmt_value(cell):
    v = cell.value
    if v is None:
        return ''
    fmt = cell.number_format or 'General'
    # Apply common Excel number formats
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Try to honor the format
        try:
            if fmt == 'General':
                # Integers stay integers
                if isinstance(v, float) and v.is_integer():
                    return f"{int(v):,}" if abs(v) >= 1000 else f"{int(v)}"
                return str(v)
            # Percentage
            if '%' in fmt:
                # decide decimals
                m = re.search(r'0\.(0+)%', fmt)
                dec = len(m.group(1)) if m else 0
                return f"{v*100:,.{dec}f}%"
            # Comma thousand separator with optional decimals
            if '#,##0' in fmt or '0,000' in fmt or ',' in fmt:
                m = re.search(r'0\.(0+)', fmt)
                dec = len(m.group(1)) if m else 0
                if v < 0 and '(' in fmt:
                    return f"({abs(v):,.{dec}f})"
                return f"{v:,.{dec}f}"
            # decimals like 0.00
            m = re.search(r'0\.(0+)', fmt)
            if m:
                dec = len(m.group(1))
                return f"{v:.{dec}f}"
            if isinstance(v, float) and v.is_integer():
                return f"{int(v):,}"
            return f"{v:,}" if abs(v) >= 1000 else str(v)
        except Exception:
            return str(v)
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%Y/%m/%d')
        except Exception:
            return str(v)
    return str(v)

def cell_classes_and_style(cell, col_w_px=None, row_h_px=None, has_value=True):
    style = []
    # Skip most styling for empty cells - keep file size sane
    if not has_value:
        # Only emit fill (so striped/background headers show through) and borders
        fill = cell.fill
        if fill and fill.fgColor and fill.patternType in ('solid',):
            bg = color_to_hex(fill.fgColor)
            if bg and bg.lower() not in ('#ffffff', '#000000'):
                style.append(f'background:{bg}')
        b = cell.border
        if b:
            for side, css in (('left','border-left'),('right','border-right'),('top','border-top'),('bottom','border-bottom')):
                s = getattr(b, side)
                if s and s.style and s.style != 'none':
                    w = '2px' if s.style in ('thick','double') else '1px'
                    kind = 'dashed' if s.style == 'dashed' else ('dotted' if s.style == 'dotted' else 'solid')
                    col = color_to_hex(s.color) or '#000'
                    style.append(f'{css}:{w} {kind} {col}')
        return ';'.join(style)
    # Font
    f = cell.font
    if f:
        if f.bold:
            style.append('font-weight:600')
        if f.italic:
            style.append('font-style:italic')
        if f.size and round(float(f.size)) != 11:
            style.append(f'font-size:{round(float(f.size))}px')
        col = color_to_hex(f.color)
        if col and col.lower() != '#000000':
            style.append(f'color:{col}')
        if f.underline:
            style.append('text-decoration:underline')
    # Fill
    fill = cell.fill
    if fill and fill.fgColor and fill.patternType in ('solid',):
        bg = color_to_hex(fill.fgColor)
        if bg and bg.lower() not in ('#ffffff', '#000000'):
            style.append(f'background:{bg}')
        elif bg and bg.lower() == '#000000':
            # Sometimes #000000 means default
            pass
    # Alignment
    a = cell.alignment
    if a:
        if a.horizontal:
            ha = {'general':'', 'left':'left', 'right':'right', 'center':'center', 'centerContinuous':'center', 'fill':'left', 'justify':'justify'}.get(a.horizontal, '')
            if ha:
                style.append(f'text-align:{ha}')
        if a.vertical:
            va = {'top':'top', 'center':'middle', 'bottom':'bottom'}.get(a.vertical, '')
            if va:
                style.append(f'vertical-align:{va}')
        if a.wrap_text:
            style.append('white-space:pre-wrap; word-break:break-word')
        else:
            style.append('white-space:nowrap')
        if a.indent:
            style.append(f'padding-left:{int(a.indent)*8 + 2}px')
    # Borders
    b = cell.border
    if b:
        for side, css in (('left','border-left'),('right','border-right'),('top','border-top'),('bottom','border-bottom')):
            s = getattr(b, side)
            if s and s.style and s.style != 'none':
                w = '1px'
                kind = 'solid'
                if s.style in ('thick','double'):
                    w = '2px'
                if s.style == 'dashed':
                    kind = 'dashed'
                if s.style == 'dotted':
                    kind = 'dotted'
                col = color_to_hex(s.color) or '#000'
                style.append(f'{css}:{w} {kind} {col}')
    return ';'.join(style)

def build_html_for_sheet(ws, title):
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    # Build merge map
    merge_origin = {}  # (r,c) -> (rowspan, colspan)
    merge_skip = set()  # cells covered by merge but not origin
    for mr in ws.merged_cells.ranges:
        rspan = mr.max_row - mr.min_row + 1
        cspan = mr.max_col - mr.min_col + 1
        merge_origin[(mr.min_row, mr.min_col)] = (rspan, cspan)
        for r in range(mr.min_row, mr.max_row+1):
            for c in range(mr.min_col, mr.max_col+1):
                if (r,c) != (mr.min_row, mr.min_col):
                    merge_skip.add((r,c))

    # Column widths
    col_widths = []
    for c in range(1, max_c+1):
        letter = get_column_letter(c)
        d = ws.column_dimensions.get(letter)
        col_widths.append(col_px(d.width if d else None))

    # Build table
    out = []
    out.append('<div class="sheet">')
    # colgroup
    out.append('<colgroup>')
    for w in col_widths:
        out.append(f'<col style="width:{w}px">')
    out.append('</colgroup>')
    out.append('<tbody>')
    for r in range(1, max_r+1):
        rd = ws.row_dimensions.get(r)
        h = rd.height if rd else None
        # Skip explicitly hidden rows
        if rd and rd.hidden:
            continue
        out.append(f'<tr style="height:{row_px(h)}px">')
        for c in range(1, max_c+1):
            if (r,c) in merge_skip:
                continue
            cell = ws.cell(row=r, column=c)
            text = fmt_value(cell)
            style = cell_classes_and_style(cell, has_value=bool(text))
            attrs = ''
            if (r,c) in merge_origin:
                rs, cs = merge_origin[(r,c)]
                if rs > 1: attrs += f' rowspan="{rs}"'
                if cs > 1: attrs += f' colspan="{cs}"'
            # newlines -> <br>
            text_html = html.escape(text).replace('\n', '<br>')
            out.append(f'<td{attrs} style="{style}">{text_html}</td>')
        out.append('</tr>')
    out.append('</tbody></div>')  # placeholder - we'll fix the wrapping below
    return '\n'.join(out), max_r, max_c

def build_html_for_sheet_table(ws):
    """Build a single <table> (no surrounding div) for a sheet."""
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    merge_origin = {}
    merge_skip = set()
    for mr in ws.merged_cells.ranges:
        rspan = mr.max_row - mr.min_row + 1
        cspan = mr.max_col - mr.min_col + 1
        merge_origin[(mr.min_row, mr.min_col)] = (rspan, cspan)
        for r in range(mr.min_row, mr.max_row+1):
            for c in range(mr.min_col, mr.max_col+1):
                if (r,c) != (mr.min_row, mr.min_col):
                    merge_skip.add((r,c))
    col_widths = []
    for c in range(1, max_c+1):
        letter = get_column_letter(c)
        d = ws.column_dimensions.get(letter)
        col_widths.append(col_px(d.width if d else None))

    out = ['<table class="xl">']
    out.append('<colgroup>')
    for w in col_widths:
        out.append(f'<col style="width:{w}px">')
    out.append('</colgroup>')
    out.append('<tbody>')
    for r in range(1, max_r+1):
        rd = ws.row_dimensions.get(r)
        h = rd.height if rd else None
        if rd and rd.hidden:
            continue
        out.append(f'<tr style="height:{row_px(h)}px">')
        for c in range(1, max_c+1):
            if (r,c) in merge_skip:
                continue
            cell = ws.cell(row=r, column=c)
            text = fmt_value(cell)
            style = cell_classes_and_style(cell, has_value=bool(text))
            attrs = ''
            if (r,c) in merge_origin:
                rs, cs = merge_origin[(r,c)]
                if rs > 1: attrs += f' rowspan="{rs}"'
                if cs > 1: attrs += f' colspan="{cs}"'
            text_html = html.escape(text).replace('\n', '<br>')
            out.append(f'<td{attrs} style="{style}">{text_html}</td>')
        out.append('</tr>')
    out.append('</tbody></table>')
    return '\n'.join(out), max_r, max_c

def main():
    wb = load_workbook(SRC, data_only=True)
    sheets_to_render = ['輸入頁', '列印頁']
    sections = []
    for name in sheets_to_render:
        ws = wb[name]
        table_html, mr, mc = build_html_for_sheet_table(ws)
        sections.append((name, table_html, mr, mc))

    # Build navigation - identify "logical pages" inside 列印頁 by rows of mostly-empty bands.
    # Also build a quick anchor list.
    page_anchors = [
        ('p1',   '頁1: 主要資料 + 彙總表(中分紅 1-41)', 1),
        ('p2',   '頁2: 彙總表(中分紅 42-90)', 60),
        ('p3',   '頁3: 彙總表(中分紅 91+)', 120),
        ('p4',   '頁4: 彙總表(較低紅利 1-50)', 180),
        ('p5',   '頁5: 彙總表(較低紅利 52-102)', 242),
        ('p6',   '頁6: 彙總表(較低紅利 103+)', 304),
        ('p7',   '頁7: 投保利益表 1-44', 366),
        ('p8',   '頁8: 投保利益表 45-87', 424),
        ('p9',   '頁9: 投保利益表 88+', 480),
        ('p10',  '頁10: 紅利說明(最可能)', 540),
        ('p11',  '頁11: 紅利說明(最可能 續)', 594),
        ('p12',  '頁12: 紅利說明(最可能 續)', 652),
        ('p13',  '頁13: 紅利說明(較低)', 710),
        ('p14',  '頁14: 紅利說明(較低 續)', 769),
        ('p15',  '頁15: 紅利說明(較低 續)', 827),
        ('p16',  '頁16: 商品摘要', 885),
        ('p17',  '頁17: 揭露事項 / 風險告知', 940),
        ('p18',  '頁18: 分紅公式說明', 999),
        ('p19',  '頁19: 簽名頁', 1039),
    ]

    # Add anchor injection by post-processing the 列印頁 HTML: insert <span id="pX"> before tr at given row
    # Easiest: re-emit with anchor markers
    wb2 = load_workbook(SRC, data_only=True)
    ws_print = wb2['列印頁']

    # Build with anchors for each page start row
    anchor_rows = {row: aid for aid,_,row in page_anchors}
    print_html = build_table_with_anchors(ws_print, anchor_rows)
    # replace 列印頁 section
    for i,(name,t,mr,mc) in enumerate(sections):
        if name == '列印頁':
            sections[i] = (name, print_html, mr, mc)

    nav_html = '<nav class="topnav"><strong>跳至：</strong> '
    nav_html += '<a href="#sheet-輸入頁">▶ 輸入頁</a> '
    nav_html += '<span class="sep">|</span> '
    for aid, label, _ in page_anchors:
        nav_html += f'<a href="#{aid}">{html.escape(label)}</a> '
    nav_html += '</nav>'

    # Final HTML
    head = """<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
<title>富邦人壽美富紅運外幣分紅終身壽險建議書</title>
<style>
:root {
  --ink: #1d1d1f;
  --bg: #f5f5f7;
  --paper: #ffffff;
  --border: #d2d2d7;
  --accent: #0a4a7a;
  --muted: #6e6e73;
}
* { box-sizing: border-box; }
html, body {
  margin: 0; padding: 0;
  background: var(--bg);
  color: var(--ink);
  font-family: 'Microsoft JhengHei', 'PingFang TC', 'Heiti TC', 'Noto Sans TC', Arial, sans-serif;
  font-size: 13px;
  -webkit-text-size-adjust: 100%;
}
.topnav {
  position: sticky; top: 0; z-index: 100;
  background: rgba(255,255,255,0.96);
  backdrop-filter: saturate(180%) blur(10px);
  border-bottom: 1px solid var(--border);
  padding: 8px 12px;
  font-size: 12px;
  line-height: 1.9;
  white-space: nowrap;
  overflow-x: auto;
}
.topnav strong { color: var(--accent); margin-right: 6px; }
.topnav a {
  color: var(--accent); text-decoration: none;
  padding: 2px 6px; border-radius: 4px;
  display: inline-block;
}
.topnav a:hover { background: #e6f0fa; }
.topnav .sep { color: var(--border); margin: 0 4px; }
.title-banner {
  background: linear-gradient(135deg, #0a4a7a, #1668a7);
  color: white;
  padding: 18px 24px;
  border-bottom: 4px solid #d4a017;
}
.title-banner h1 {
  margin: 0 0 4px 0; font-size: 22px; font-weight: 700;
  font-family: 'Microsoft JhengHei', 'PingFang TC', sans-serif;
}
.title-banner .sub { font-size: 12px; opacity: 0.9; }
.section {
  background: var(--paper);
  margin: 16px auto;
  max-width: 1100px;
  padding: 0;
  border: 1px solid var(--border);
  border-radius: 8px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.04);
  overflow: hidden;
}
.section h2 {
  margin: 0;
  padding: 10px 16px;
  background: #fafafa;
  border-bottom: 1px solid var(--border);
  font-size: 14px;
  color: var(--accent);
}
.scroll {
  overflow: auto;
  max-height: none;
}
table.xl {
  border-collapse: collapse;
  table-layout: fixed;
  width: max-content;
  font-size: 12px;
  background: white;
}
table.xl td {
  padding: 1px 3px;
  vertical-align: middle;
  overflow: hidden;
  font-family: inherit;
}
table.xl td:empty { background: transparent; }
.anchor-marker { display: block; height: 0; visibility: hidden; }
@media print {
  .topnav, .title-banner { display: none; }
  .section { box-shadow: none; border: none; margin: 0; }
}
@media (max-width: 720px) {
  .section { margin: 8px 0; border-radius: 0; max-width: 100%; }
  body { font-size: 12px; }
}
</style>
</head>
<body>
<div class="title-banner">
  <h1>富邦人壽美富紅運外幣分紅終身壽險建議書</h1>
  <div class="sub">PFA · V2.6-1150413 · 試算日 1150425（網頁版）</div>
</div>
"""
    body = [head, nav_html]
    for name, table_html, mr, mc in sections:
        slug = 'sheet-' + name
        body.append(f'<section class="section" id="{slug}">')
        body.append(f'<h2>{html.escape(name)}（{mr} × {mc}）</h2>')
        body.append('<div class="scroll">')
        body.append(table_html)
        body.append('</div></section>')
    body.append('</body></html>')

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(body))
    print(f'Wrote {OUT}')

def build_table_with_anchors(ws, anchor_rows):
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    merge_origin = {}
    merge_skip = set()
    for mr in ws.merged_cells.ranges:
        rspan = mr.max_row - mr.min_row + 1
        cspan = mr.max_col - mr.min_col + 1
        merge_origin[(mr.min_row, mr.min_col)] = (rspan, cspan)
        for r in range(mr.min_row, mr.max_row+1):
            for c in range(mr.min_col, mr.max_col+1):
                if (r,c) != (mr.min_row, mr.min_col):
                    merge_skip.add((r,c))
    col_widths = []
    for c in range(1, max_c+1):
        letter = get_column_letter(c)
        d = ws.column_dimensions.get(letter)
        col_widths.append(col_px(d.width if d else None))

    out = ['<table class="xl">']
    out.append('<colgroup>')
    for w in col_widths:
        out.append(f'<col style="width:{w}px">')
    out.append('</colgroup>')
    out.append('<tbody>')
    for r in range(1, max_r+1):
        rd = ws.row_dimensions.get(r)
        h = rd.height if rd else None
        if rd and rd.hidden:
            continue
        anchor = ''
        if r in anchor_rows:
            anchor = f'<a id="{anchor_rows[r]}" class="anchor-marker"></a>'
        out.append(f'<tr style="height:{row_px(h)}px">')
        first_emitted = False
        for c in range(1, max_c+1):
            if (r,c) in merge_skip:
                continue
            cell = ws.cell(row=r, column=c)
            text = fmt_value(cell)
            style = cell_classes_and_style(cell, has_value=bool(text))
            attrs = ''
            if (r,c) in merge_origin:
                rs, cs = merge_origin[(r,c)]
                if rs > 1: attrs += f' rowspan="{rs}"'
                if cs > 1: attrs += f' colspan="{cs}"'
            text_html = html.escape(text).replace('\n', '<br>')
            cell_inner = (anchor + text_html) if (anchor and not first_emitted) else text_html
            out.append(f'<td{attrs} style="{style}">{cell_inner}</td>')
            first_emitted = True
        # If row had no cell emitted at all and we have an anchor, emit a placeholder td
        out.append('</tr>')
    out.append('</tbody></table>')
    return '\n'.join(out)

if __name__ == '__main__':
    main()
